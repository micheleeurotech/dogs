#authors Michele Della Mea, Angelo Calabria
#We are two trainee and we made this project between february and april 2016 for Eurotech Ltd. (Cambridge, UK) to compare the prices of the electronic components 
#provided by Radiance and Octopart. This application is tested by python 3.5. The radiance prices are located in a xlsx file and the prices of Octopart are located in a
#csv file downloaded from octopart.com . To obtain the Octopart file we must upload the Radiance file in the web site clicking the bottom "UPLOAD BOM" giving to
#the web site the exact column-location of the mpn and of the quantity. So we will obtain a csv file with the Octopart prices. By this application you  
#compare the prices of Octopart and the prices contained in the BOM of Radiance obtaining a new xlsx file that analizes the prices and you will see which 
#supplier offer the best prices. Obviously the Radiance BOM must follow a standard model otherwise it'll be impossible to compare the prices. If you want to build 
#the std BOM model and to have further information we will invite you at reading the txt file containing every information that you need.

#libraries that we are going to use
from tkinter import *   
from tkinter.filedialog import askopenfilename
import pandas as pd
import numpy as np
import openpyxl
import os.path, time
import datetime
import shutil
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Font
from openpyxl.cell import get_column_letter
import webbrowser


#these variables are going to contain the paths of the files and the board name
radiance_Path = ""
octopart_Path = ""
result_Path = ""
board_name = ""

class Application:      
	
	def __init__(self, root):  
		
		#insert radiance path file by the wizard
		def callback():
			name= askopenfilename()
			radiance_entry.insert(0,name)           
			radiance_Path = name
	
		#insert octopart path file by the wizard
		def callback2():
			name = askopenfilename()
			octopart_entry.insert(0,name)
			octopart_Path = name

		#opening octopart web site by the wizard
		def callback3(event):
			webbrowser.open_new(r"https://octopart.com/")

		#this long method executes the program
		def execute():
			#saving variables that you have insert by the wizard
			radiance_Path = radiance_entry.get()
			octopart_Path = octopart_entry.get()
			result_Name = result_entry.get()
			board_name = bName_entry.get()
			
			#changing format from .XLSX to .xlsx
			if "XLSX" in radiance_Path:
				radiance_Path = radiance_Path.replace("XLSX", "xlsx")

			#checking correctness excel file format (.xlsx)
			if radiance_Path[-5:] != ".xlsx":
				text_exception = "Pay attention because the format of the excel file is not .xlsx"
				Error.configure(text=text_exception)

			#assigning paths of the three files (2 sources and 1 destination) and number board
			board = board_name
			from_rad = radiance_Path #excel
			from_octo = octopart_Path #csv
			to_merge = result_Name #excel

			#checking correctness destination excel file format 
			if to_merge[-5:] != ".xlsx":
				text_exception = "Pay attention because the format of the resulted excel file is not .xlsx"
				Error.configure(text=text_exception)

			#extracting board size from the octopart column that is inserted between the two following variables, 
			#for example:"Unit Price (@ Batch Size: 5000) [Lowest Price (Overall)]"
			pre_size = "Unit Price (@ Batch Size: "
			post_size = ") [Lowest Price (Overall)]"
			
			#assigning colname with octopart prices to a variable
			distributor_octo = 'Distributor [Lowest Price (Overall)]'
			
			#reading octopart csv file and checking the correctness of the format. As long as it's incorrect printing
			#message on the wizard
			try:
				df1 = pd.read_csv(from_octo)
			except:
				text_exception = "Pay attention because the format of the csv file is not .csv"
				Error.configure(text=text_exception)

			#extracting board size from octopart using the variables pre_size and post_size
			for a in df1.columns:
				if(pre_size in a):
					if(post_size in a):
						size_price = a

			temp = size_price.split()
			batch_size = temp[5].rstrip(")")
			
			#creating data frame df1 using the three columns that we need from octopart
			df1 = df1[['Line Item', distributor_octo, size_price]]
			
			#deleting unuseful rows
			df1 = df1.iloc[3:,:]
			
			#renaming columns
			df1.columns = ['MPN', 'Lowest MFR Octo', 'Price Octo']
			
			#reading radiance excel file
			xl = pd.ExcelFile(from_rad)
			
			#examining sheet 1 of the previous file
			sheet1 = xl.parse(0)

			#create new copy file
			shutil.copy(from_rad, to_merge)
			
			#selecting useful columns from radiance file and assigning them to variables
			mfr = "MFR1"
			u_price = "Unit Price (US$)"
			mpn = "MPN1"
			rem = "Remarks"
			eur_ref = "Eurotech Ref"
			
			#creating variables that we are going to use to save key-cell positions
			cellA = 0
			name_board = 0
			acc = 0
			
			#finding eurotech ref cell position because in this way we will know where exactly the data frame starts
			for i in range(0, len(sheet1.iloc[:,0])):
				if sheet1.iloc[i,0] == eur_ref:
					cellA = i
					break

			#checking the presence of a cell called eurotech ref on radiance, otherwise printing message error on the wizard
			if sheet1.iloc[cellA,0] != eur_ref:
				text_exception = "Pay attention because there is not a 'Eurotech Ref' column in the file"
				Error.configure(text=text_exception)
			
			#creating variables that we are going to use to save key-cell positions
			cellB = 0
			cellC = 0
			cellD = 0
			cellF = 0
			cellE = 0
			
			#creating df2 subsetting the sheet 1 of radiance starting from the cell eurotech ref
			df2 = sheet1.iloc[cellA:,:]
			
			acc2 = 0

			#finding key-cells column positions in radiance
			for i in range(0, len(sheet1.iloc[0,:])):
				if df2.iloc[0,i] == mpn:
					cellB = i
				elif df2.iloc[0,i] == u_price:  
					cellC = i
				elif df2.iloc[0,i] == board:
					if acc2 == 0:   
						cellD = i
						acc2 = 1
				elif df2.iloc[0,i] == mfr:
					cellE = i
				elif df2.iloc[0,i] == rem:
					cellF = i
				else: 
					continue

			#checking missing columns in radiance and printing message on the wizard containing missing colnames
			if df2.iloc[0,cellB] != mpn or df2.iloc[0,cellC] != u_price or df2.iloc[0,cellD] != board or df2.iloc[0,cellE] != mfr or df2.iloc[0,cellF] != rem:
				dictionary = {mpn: cellB, u_price: cellC, board: cellD, mfr: cellE, rem: cellF}
				listDict = []
				for key in dictionary:
					if dictionary[key] == 0:
						listDict.append(key)
				text_exception = "Pay attention because the following columns are missing: " + ', '.join(listDict)
				Error.configure(text=text_exception)
			
			#assigning to df2 the values on radiance contained in the key columns
			df2 = df2.iloc[1:,[0,cellE,cellB,cellC,cellD]]  
			
			#renaming columns df2
			df2.columns = ['Eurotech Ref', 'MFR', 'MPN', 'Price Rad', 'Qty']
			
			#excluding rows with MPN null
			df2 = df2[df2.MPN.notnull()]
			
			#creating df3 merging df1 and df2
			df3 = pd.merge(df1,df2,on='MPN',how='right')
			
			#moving column eurotech ref to index
			df3.set_index('Eurotech Ref', inplace=True)
			
			#replacing 0 with nan
			df3.replace(0, np.nan, inplace=True)
			
			#filling in new columns containing the difference of the prices or nan
			df3['Octo - Rad'] = df3['Price Octo'] - df3['Price Rad']
			df3["Best Sub"] = np.nan
			df3["Best Price"] = np.nan
			
			#positioning colnames
			cols = ['MFR', 'MPN', 'Qty', 'Price Rad', 'Price Octo', 'Octo - Rad', "Best Sub", "Lowest MFR Octo", "Best Price"]
			
			df3 = df3[cols]
			
			#filling in Best Price and Best Sub columns. Remember that the current columns order is contained in the list "cols" (as above) so
			#it'll be easier to understand the code
			for i in range(0, len(df3.iloc[:,0])):
				if df3.iloc[i,5] >= 0:
					df3.iloc[i,8] = df3.iloc[i,3]
					df3.iloc[i,6] = "Rad"
				elif df3.iloc[i,5] < 0:
					df3.iloc[i,8] = df3.iloc[i,4]
					df3.iloc[i,6] = "Octo"
				elif pd.isnull(df3.iloc[i,4]) and pd.notnull(df3.iloc[i,3]):
					df3.iloc[i,8] = df3.iloc[i,3]
					df3.iloc[i,6] = "Rad"
				elif pd.notnull(df3.iloc[i,4]) and pd.isnull(df3.iloc[i,3]):
					df3.iloc[i,8] = df3.iloc[i,4]
					df3.iloc[i,6] = "Octo"
				else:
					continue   
			
			#initialise new variable that is going to contain the cost saving using the best prices comparing octopart and radiance
			cost_saving = 0
			
			#calculating cost saving using columns 'Qty' and 'Octo - Rad'. Remember that he current order of the colnames of df3 
			#is the list "cols"
			for i in range(0, len(df3.iloc[:,0])):
				if pd.isnull(df3.iloc[i,3]):
					df3.iloc[i,7] = np.nan
				elif df3.iloc[i,6] == "Rad":
					df3.iloc[i,7] = np.nan
				elif pd.notnull(df3.iloc[i,7]) & pd.notnull(df3.iloc[i,2]):
					cost_saving += df3.iloc[i,2] * df3.iloc[i,5]
				else:
					continue
			
			#absolute value
			cost_saving = abs(cost_saving)
			
			#create new column in df3
			df3['P * Qty'] = df3['Best Price'] * df3['Qty'] 
			
			#adding new row to df3 with rounded totals
			df3.loc['Total']= round(df3.sum(), 4)
			
			#finding total cost of the board with best prices value
			total_cost = df3.iloc[len(df3.index)-1, len(df3.columns)-1]

			#extracting octopart datetime from octopart file
			time_now = os.path.getmtime(from_octo)
			time_now = datetime.datetime.fromtimestamp(time_now)
			
			#open workbook of the destination file in which we are going to write. ws in the first sheet of the destination file
			yfile = openpyxl.load_workbook(to_merge, data_only=True)
			ws = yfile.worksheets[0]

			#creating new sheet in the destination file
			sheet2 = yfile.create_sheet()

			#assigning name to ws2 that is the second sheet of the destination file
			ws2 = yfile.worksheets[len(yfile.get_sheet_names())-1].title = board
			ws2 = yfile.worksheets[len(yfile.get_sheet_names())-1]

			#creating patterns to color excel files and modify cell formats
			colorFill = PatternFill(start_color='FFFFFF00', #yellow
                   fill_type='solid')

			font = Font(name='Calibri',
                 size=11,
                 bold=False,
                 color='FF000000') #black
			
			#coloring and formatting the cell A1 in the second sheet
			ws2.cell(row=1, column = 1).value = "Eurotech Ref"
			ws2.cell(row=1, column = 1).fill = colorFill
			ws2.cell(row=1, column = 1).font = Font(bold=True)

			#writing df3 columns in the second sheet
			for i in range (1, len(df3.columns)+1):
					ws2.cell(row=1, column = i+1).value = df3.columns[i-1]
					ws2.cell(row=1, column = i+1).fill = colorFill
					ws2.cell(row=1, column = i+1).font = Font(bold=True)

			#writing index in the second sheet (if you forget what the sheets refer remember that we are writing only
			#in the destination file and there are only two sheets in which this program writes)
			for j in range (1, len(df3.index)+1):
				ws2.cell(row=j+1, column = 1).value = df3.index[j-1]
				ws2.cell(row=j+1, column = 1).font = font

			#writing all the values contained in df3 on the second sheet of the destination file
			for i in range (1, len(df3.iloc[1,:])+1):
				for j in range (1, len(df3.iloc[:,1])+1):
					if pd.notnull(df3.iloc[j-1,i-1]):
						ws2.cell(row=j+1, column = i+1).value = df3.iloc[j-1,i-1]
						ws2.cell(row=j+1, column = i+1).font = font
					if pd.notnull(df3.iloc[j-1,i-1]) and (i==2 or i==7 or i==10):
						ws2.cell(row=j+1, column = i+1).font = Font(bold=True)

			#creating variables that are going to contain the total board cost using exclusively octopart or radiance
			#prices (when they exist...)
			tot_rad = 0
			tot_octo = 0

			#calculating totals from df3 using the two variables just created
			for i in range(0, len(df3.iloc[:,3])-2):
				if pd.notnull(df3.iloc[i,3]) and pd.notnull(df3.iloc[i,2]):
					tot_rad += df3.iloc[i,3] * df3.iloc[i,2]
				if pd.notnull(df3.iloc[i,4]) and pd.notnull(df3.iloc[i,2]):
					tot_octo += df3.iloc[i,4] * df3.iloc[i,2]

			#creating variable containing the percentage saving
			sav100 = (cost_saving/tot_rad * 100)

			#creating variable that is going to contain eurotech ref row in the first sheet
			eur_row = 0

			#finding eurotech ref position to know when the data frame starts in the first sheet
			while(True):
				eur_row+=1
				if ws.cell(row=eur_row, column = 1).value == eur_ref:
					break
				if (acc==len(sheet1.iloc[:,0])):
					break

			#writing, styling and adding new colnames in the first sheet
			ws.cell(row=eur_row, column = cellF+4).value = "Low MFR Octo"
			ws.cell(row=eur_row, column = cellF+3).value = "Low Price Octo"
			ws.cell(row=eur_row, column = cellF+2).value = "Price Rad"
			ws.cell(row=eur_row, column = cellF+5).value = "Diff"
			ws.cell(row=eur_row, column = cellF+6).value = "P * Qty"

			for i in range(2,7):
				ws.cell(row=eur_row, column = cellF+i).fill = colorFill
				ws.cell(row=eur_row, column = cellF+i).font = Font(bold=True)

			##writing, styling and adding new fixed cells in the first sheet
			ws.cell(row=eur_row, column = cellF+8).value = "Saving"
			ws.cell(row=eur_row+1, column = cellF+8).value = "Saving %"
			ws.cell(row=eur_row+2, column = cellF+8).value = "Tot Board Rad"
			ws.cell(row=eur_row+3, column = cellF+8).value = "Tot Board Octo"
			ws.cell(row=eur_row+4, column = cellF+8).value = "Batch Size"
			ws.cell(row=eur_row+5, column = cellF+8).value = "Date Time"

			for i in range(0,6):
				ws.cell(row=eur_row+i, column = cellF+8).fill = colorFill 
				ws.cell(row=eur_row+i, column = cellF+8).font = Font(bold=True)

			##writing, styling and adding values on the cells on the right of the previous cells in the first sheet
			ws.cell(row=eur_row, column = cellF+9).value = round(cost_saving, 4)
			ws.cell(row=eur_row+1, column = cellF+9).value = round(sav100, 4)
			ws.cell(row=eur_row+2, column = cellF+9).value = round(tot_rad, 4)
			ws.cell(row=eur_row+3, column = cellF+9).value = round(tot_octo, 4)
			ws.cell(row=eur_row+4, column = cellF+9).value = batch_size
			ws.cell(row=eur_row+5, column = cellF+9).value = time_now

			for i in range(0,6):
				ws.cell(row=eur_row+i, column = cellF+9).font = Font(bold=True)

			#creating variable that is going to contain the carrier board name
			board_position = 0

			#finding name board cell in the first sheet and the carrier board name
			for i in range(1, len(df3.iloc[:,0])):
				if ws.cell(row=i, column = 1).value == board:
					name_board = ws.cell(row=i, column = 2).value
					break

			#print(df3.columns)

			#filling in the new columns previously created in the first sheet using the data contained in df3. The order of df3 is the
			#following: ['MFR', 'MPN', 'Qty', 'Price Rad', 'Price Octo', 'Octo - Rad','Best Sub', 'Lowest MFR Octo', 'Best Price', 'P * Qty']
			for i in range(1, len(sheet1.iloc[:,0])):
				if pd.notnull(ws.cell(row=eur_row+i, column = cellB).value):
					if pd.notnull(ws.cell(row=eur_row+i, column = 1).value):
						for j in range(0, len(df3.iloc[:,0])-1):
							if pd.notnull(df3.iloc[j,2]):
								if (ws.cell(row=eur_row+i, column = 1).value)==df3.index[j]:
									if pd.notnull(df3.iloc[j,7]):
										ws.cell(row=eur_row+i, column = cellF+4).value = df3.iloc[j,7]
										ws.cell(row=eur_row+i, column = cellF+3).value = df3.iloc[j,4]
										ws.cell(row=eur_row+i, column = cellF+4).font = font
										ws.cell(row=eur_row+i, column = cellF+3).font = font
										if pd.notnull(df3.iloc[j,5]):
											ws.cell(row=eur_row+i, column = cellF+5).value = abs(df3.iloc[j,5])
											ws.cell(row=eur_row+i, column = cellF+5).font = font
									ws.cell(row=eur_row+i, column = cellF+2).value = df3.iloc[j,3]
									ws.cell(row=eur_row+i, column = cellF+2).font = font
									ws.cell(row=eur_row+i, column = cellF+6).value = df3.iloc[j,9]
									ws.cell(row=eur_row+i, column = cellF+6).font = Font(bold=True)

			#filling in missing values in the first sheet. Sometimes there are radiance components without MPN that we haven't 
			#examined previously. In the first sheet we write also those components highlighting them with yellow cells. df3 columns
			#order is as above
			for i in range(1, len(sheet1.iloc[:,0])):
				if pd.notnull(ws.cell(row=eur_row+i, column = cellC+1).value):
					if pd.notnull(ws.cell(row=eur_row+i, column = cellD+1).value):
						if pd.isnull(ws.cell(row=eur_row+i, column = cellF+2).value):
							ws.cell(row=eur_row+i, column = cellF+6).value = float(ws.cell(row=eur_row+i, column = cellD+1).value) * float(ws.cell(row=eur_row+i, column = cellC+1).value)
							ws.cell(row=eur_row+i, column = cellF+6).font = Font(bold=True)
							ws.cell(row=eur_row+i, column = cellF+2).value = ws.cell(row=eur_row+i, column = cellC+1).value
							for j in range(2,7):
								ws.cell(row=eur_row+i, column=cellF+j).fill = colorFill
								ws.cell(row=eur_row+i, column=cellF+j).font = font
								if j == 6:
									ws.cell(row=eur_row+i, column=cellF+j).fill = colorFill
									ws.cell(row=eur_row+i, column=cellF+j).font = Font(bold=True)
			
			#creating variables that are going to find the total cost of the board position and its value
			total_position = 0
			total_value = 0

			#writing and finding total board cost cell in the first sheet and calculating its value. The total board cost
			#will be located two cells below the last value of the last new column ("Qty*P")
			for i in range(1, len(sheet1.iloc[:,0])):
				if pd.notnull(ws.cell(row=eur_row+i, column = cellF+6).value):
					total_position = i
					total_value += float(ws.cell(row=eur_row+i, column = cellF+6).value)

			#styling, assigning and formatting total board cost cell in the first sheet
			ws.cell(row=total_position+eur_row+2, column = cellF+6).value = round(total_value, 4)
			ws.cell(row=total_position+eur_row+2, column = cellF+6).font = Font(bold=True)
			ws.cell(row=total_position+eur_row+2, column = cellF+6).fill = colorFill

			#writing and styling second sheet fixed cells
			ws2.cell(row=1, column=len(df3.columns)+3).value = "Board"
			ws2.cell(row=2, column=len(df3.columns)+3).value = "Date"
			ws2.cell(row=3, column=len(df3.columns)+3).value = "Cost Saving"
			ws2.cell(row=4, column=len(df3.columns)+3).value = "Total Cost"

			for i in range(1,5):
				ws2.cell(row=i, column=len(df3.columns)+3).fill = colorFill
				ws2.cell(row=i, column=len(df3.columns)+3).font = Font(bold=True)

			#assigning values and styling the cells just created in the second sheet
			ws2.cell(row=1, column=len(df3.columns)+4).value = name_board
			ws2.cell(row=2, column=len(df3.columns)+4).value = time_now
			ws2.cell(row=3, column=len(df3.columns)+4).value = round(cost_saving, 4)
			ws2.cell(row=4, column=len(df3.columns)+4).value = round(total_cost, 4)
			
			for i in range(1,5):
				ws2.cell(row=i, column=len(df3.columns)+4).font = Font(bold=True)

			#styling, writing and formatting the last row of the data frame in the second sheet			
			ws2.cell(row=len(df3.index)+1, column = 1).fill = colorFill
			ws2.cell(row=len(df3.index)+1, column = 1).font = Font(bold=True)

			for i in range(1,len(df3.iloc[1,:])+1):
				if pd.notnull(df3.iloc[len(df3.index)-1,i-1]):
					ws2.cell(row=len(df3.index)+1, column = i+1).fill = colorFill
					ws2.cell(row=len(df3.index)+1, column = i+1).font = Font(bold=True)
				else:
					ws2.cell(row=len(df3.index)+1, column = i+1).value = "-"
					ws2.cell(row=len(df3.index)+1, column = i+1).fill = colorFill
					ws2.cell(row=len(df3.index)+1, column = i+1).font = Font(bold=True)

			#excel cells resizing in both sheets. For example when we know that there is a timedate in a cell, we
			#are going to enlarge that cell
			ws2.column_dimensions["A"].width = 14
			ws2.column_dimensions["C"].width = 26
			ws2.column_dimensions["D"].width = 5
			ws2.column_dimensions["H"].width = 10
			ws2.column_dimensions["I"].width = 18
			ws2.column_dimensions["M"].width = 10
			ws2.column_dimensions["N"].width = 20

			ws.column_dimensions[get_column_letter(cellF+2)].width = 14
			ws.column_dimensions[get_column_letter(cellF+3)].width = 14
			ws.column_dimensions[get_column_letter(cellF+4)].width = 16
			ws.column_dimensions[get_column_letter(cellF+5)].width = 14
			ws.column_dimensions[get_column_letter(cellF+6)].width = 14
			ws.column_dimensions[get_column_letter(cellF+8)].width = 17
			ws.column_dimensions[get_column_letter(cellF+9)].width = 19
			
			#we save all the data in one .xlsx file
			yfile.save(to_merge)

			#if the program ends without problems the following message will appear in the wizard
			Error.configure(text='Ok the program is complete')
				
		#assigning tkinter object root to the class Application and giving title
		self.root = root  
		self.root.title('Analyse your Excel File')  
		
		#creating frame 
		Frame(self.root, width=600, height=400).pack()  

		#creating label on the wizard
		Label(self.root, text='Select the two file path and click on execute to run the application').place(x=10, y=10) 
		
		#creating buttons with actions
		Button(text='Radiance Path', command=callback).place(x=10,y=50)
		Button(text='Octopart Path', command=callback2).place(x=10,y=150)
		Button(text='Execute', command=execute).place(x=10,y=300)
		
		#creating octopart path and radiance path entries on the wizard
		radiance_entry = Entry(self.root,width=60)
		radiance_entry.place(x=180,y=50)
		
		octopart_entry = Entry(self.root,width=60)
		octopart_entry.place(x=180,y=150)
		
		#creating label on the wizard
		Label(self.root, text="Insert new file Name exp.xlsx").place(x=10,y=200)
		result_entry = Entry(self.root,width=60)
		result_entry.place(x=180,y=200)
		
		Label(self.root, text="Insert the board column name").place(x=10,y=250)
		bName_entry = Entry(self.root,width=60)
		bName_entry.place(x=180,y=250)

		#creating hyperlink on the wizard
		link = Label(self.root, text="Octopart Hyperlink", fg="blue", cursor="hand2")
		link.pack()
		link.bind("<Button-1>", callback3)
		link.place(x=10, y=100)
		Label(self.root, text="Click the link for the Octopart web site").place(x=180,y=100)

		#creating error management label
		text_exception = ""
		Error = Label(self.root, text=text_exception)
		Error.pack()
		Error.place(x=80, y=300)
		
		
#creating tkinter object	
root = Tk()  
Application(root)  
root.mainloop()  

